import openpyxl
import datetime
from pygame import mixer
import time


class FixedAlarm:

    def __init__(self, alarm_name, status, alarm_time, monday, tuesday, wednesday, thursday, friday, saturday, sunday,
                 music_name):
        self.alarm_name = alarm_name
        self.status = status
        self.alarm_time = alarm_time
        self.monday = monday
        self.tuesday = tuesday
        self.wednesday = wednesday
        self.thursday = thursday
        self.friday = friday
        self.saturday = saturday
        self.sunday = sunday
        self.music_name = music_name


def read_excel():
    global alarms
    alarms = []

    file_path = "C:\\school_alarms\\school_alarms.xlsx"

    wb = openpyxl.load_workbook(file_path)
    ws = wb["Sheet1"]

    max_row = ws.max_row

    for i in range(3, max_row + 1):
        for j in range(1, 12):
            if j == 1:
                alarm_name = ws.cell(row=i, column=j).value

            elif j == 2:
                status = ws.cell(row=i, column=j).value.lower()

            elif j == 3:
                try:
                    alarm_time = int(ws.cell(row=i, column=j).value)
                except:
                    alarm_time = 0

            elif j == 4:
                monday = ws.cell(row=i, column=j).value.lower()
            elif j == 5:
                tuesday = ws.cell(row=i, column=j).value.lower()
            elif j == 6:
                wednesday = ws.cell(row=i, column=j).value.lower()
            elif j == 7:
                thursday = ws.cell(row=i, column=j).value.lower()
            elif j == 8:
                friday = ws.cell(row=i, column=j).value.lower()
            elif j == 9:
                saturday = ws.cell(row=i, column=j).value.lower()
            elif j == 10:
                sunday = ws.cell(row=i, column=j).value.lower()

            elif j == 11:
                music_name = ws.cell(row=i, column=j).value.lower()

        alarms.append(
            FixedAlarm(alarm_name, status, alarm_time, monday, tuesday, wednesday, thursday, friday, saturday, sunday,
                       music_name))


def get_current_time():
    current_time = datetime.datetime.now()
    return (current_time.hour * 100) + current_time.minute


def get_weekday():
    current_weekday = datetime.datetime.today().weekday()
    return current_weekday


def check_alarm():
    # global alarm_status, alarm_id
    # alarm_status = False
    # alarm_id = 0

    current_time = get_current_time()

    for i in range(0, len(alarms)):
        if alarms[i].alarm_time == current_time:
            alarm_status = True
            alarm_id = i
            break
        else:
            alarm_status = False
            alarm_id = None

    return current_time, alarm_status, alarm_id


def loop():
    global previous_time
    previous_time = 0
    # alarm_id =

    while 1:
        return_list = check_alarm()
        current_time = return_list[0]
        alarm_status = return_list[1]
        alarm_id = return_list[2]

        play_alarm = False

        if not (current_time == previous_time):
            previous_time = current_time
            read_excel()
            time.sleep(0.2)

            if (alarm_status):
                day_val = get_weekday()
                if day_val == 0:
                    if alarms[alarm_id].monday == "on":
                        play_alarm = True
                elif day_val == 1:
                    if alarms[alarm_id].tuesday == "on":
                        play_alarm = True
                elif day_val == 2:
                    if alarms[alarm_id].wednesday == "on":
                        play_alarm = True
                elif day_val == 3:
                    if alarms[alarm_id].thursday == "on":
                        play_alarm = True
                elif day_val == 4:
                    if alarms[alarm_id].friday == "on":
                        play_alarm = True
                elif day_val == 5:
                    if alarms[alarm_id].saturday == "on":
                        play_alarm = True
                elif day_val == 6:
                    if alarms[alarm_id].sunday == "on":
                        play_alarm = True

                if play_alarm:
                    music_path = "C:\\school_alarms\\music\\" + alarms[alarm_id].music_name + ".mp3"
                    mixer.init()
                    mixer.music.load(music_path)
                    mixer.music.play(1)
                    print("playing", music_path, current_time)

                    while(mixer.music.get_busy()):
                        time.sleep(1)

                    print("played")

        time.sleep(1)


if __name__ == '__main__':
    read_excel()
    loop()
